"""Import manually reviewed external example tables into validation draft CSVs."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from pydantic import ValidationError

from denbust.news_items.normalize import canonicalize_news_url
from denbust.taxonomy import default_taxonomy
from denbust.validation.collect import serialize_draft_rows
from denbust.validation.common import DRAFT_COLUMNS, write_csv_rows
from denbust.validation.dataset import validate_reviewed_row
from denbust.validation.models import ValidationDraftRow

TFHT_MANUAL_TRACKING_V1 = "tfht_manual_tracking_v1"
VALIDATION_REVIEWED_EXAMPLES_V1 = "validation_reviewed_examples_v1"
_MONTH_NAMES = {
    "ינואר": 1,
    "פברואר": 2,
    "מרץ": 3,
    "אפריל": 4,
    "מאי": 5,
    "יוני": 6,
    "יולי": 7,
    "אוגוסט": 8,
    "ספטמבר": 9,
    "אוקטובר": 10,
    "נובמבר": 11,
    "דצמבר": 12,
}


@dataclass(frozen=True)
class ReviewedTableImportResult:
    """Summary of a reviewed-table import run."""

    output_path: Path
    imported_rows: int
    skipped_rows: int
    warnings: list[str]


def _default_output_path(input_path: Path, format_name: str) -> Path:
    return input_path.with_name(f"{input_path.stem}_{format_name}_draft.csv")


def _extract_urls(value: str) -> list[str]:
    return re.findall(r"https?://[^\s;,\]]+", value)


def _source_name_for_url(url: str) -> str:
    netloc = urlsplit(url).netloc.lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    return netloc.split(".")[0]


def _parse_day(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    match = re.match(r"(\d{1,2})", text)
    if not match:
        return None
    return int(match.group(1))


def _infer_city(address: str) -> str:
    parts = [part.strip() for part in re.split(r"[,/]", address) if part.strip()]
    return parts[-1] if parts else ""


def _infer_taxonomy_leaf(
    event_label: str, relevant_details: str, status_text: str
) -> tuple[str, str] | None:
    combined = " ".join([event_label, relevant_details, status_text]).casefold()
    if "ערעור" in combined and "צו" in combined:
        return "brothels", "closure_appeal"
    if "סגירה מנהלית" in combined or "צו מנהלי" in combined or "סגור עד" in combined:
        return "brothels", "administrative_closure"
    if "סחר" in combined and ("בני אדם" in combined or "נשים" in combined):
        if "עבדות" in combined:
            return "human_trafficking", "trafficking_slavery_conditions"
        if any(token in combined for token in ("ברזיל", 'מחו"ל', "ייבוא", "הוסגר", "סרביה")):
            return "human_trafficking", "trafficking_cross_border_prostitution"
        if "נשים" in combined:
            return "human_trafficking", "trafficking_women"
        return "human_trafficking", "trafficking_sexual_exploitation"
    if "כתב אישום" in combined and any(
        token in combined for token in ("בית בושת", "החזקת מקום", "השכרת מקום")
    ):
        return "brothels", "brothel_indictment"
    if "קנס" in combined and "זנאי" in combined:
        return "brothels", "client_fine"
    if "שידול לזנות" in combined:
        return "pimping_prostitution", "soliciting_prostitution"
    if any(token in combined for token in ("הביא אותם לעיסוק בזנות", "הבאת אדם לידי זנות")):
        return "pimping_prostitution", "bringing_into_prostitution"
    if any(token in combined for token in ("סרסור", "סרסרות")):
        return "pimping_prostitution", "pimping"
    if "זנות מקוונת" in combined:
        return "pimping_prostitution", "online_prostitution"
    if "השכירה" in combined or "השכרת מקום" in combined:
        return "brothels", "renting_brothel"
    if "פרסום" in combined and "זנות" in combined:
        return "brothels", "advertising_prostitution"
    if any(token in combined for token in ("בית בושת", "מקום לשם זנות", "דירה דיסקרטית")):
        return "brothels", "keeping_brothel"
    return None


def _normalize_headers(row: dict[str, object]) -> dict[str, str]:
    return {
        re.sub(r"[^a-z0-9]+", "_", str(key).strip().casefold()).strip("_"): (
            "" if value is None else str(value).strip()
        )
        for key, value in row.items()
        if key is not None
    }


def _string_value(normalized: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = normalized.get(key, "").strip()
        if value:
            return value
    return ""


def _load_workbook(input_path: Path) -> Any:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    return load_workbook(input_path, data_only=True)


def _load_workbook_values(input_path: Path) -> list[tuple[object, ...]]:
    workbook = _load_workbook(input_path)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def _read_tabular_rows(input_path: Path) -> list[dict[str, object]]:
    suffix = input_path.suffix.casefold()
    if suffix == ".csv":
        import csv

        with input_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            return [dict(row) for row in reader]

    if suffix in {".xlsx", ".xlsm"}:
        rows = _load_workbook_values(input_path)
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        parsed_rows: list[dict[str, object]] = []
        for row in rows[1:]:
            parsed_rows.append(dict(zip(headers, row, strict=False)))
        return parsed_rows

    raise ValueError(f"Unsupported reviewed examples file type: {input_path.suffix}")


def _normalize_reviewed_examples_row(
    row: dict[str, object],
    *,
    collected_at: datetime,
) -> dict[str, str]:
    normalized = _normalize_headers(row)
    url = _string_value(normalized, "url", "source_url", "article_url")
    if not url:
        raise ValueError("url is required")
    article_date = _string_value(
        normalized,
        "article_date",
        "event_date",
        "publication_datetime",
        "date",
    )
    if not article_date:
        raise ValueError("article_date is required")
    title = _string_value(normalized, "title", "headline")
    if not title:
        raise ValueError("title is required")

    canonical_url = canonicalize_news_url(_string_value(normalized, "canonical_url") or url)
    source_name = _string_value(normalized, "source_name") or _source_name_for_url(url)
    snippet = _string_value(normalized, "snippet", "summary", "description")

    draft_row = {
        "source_name": source_name,
        "article_date": article_date,
        "url": url,
        "canonical_url": canonical_url,
        "title": title,
        "snippet": snippet,
        "suggested_relevant": _string_value(normalized, "suggested_relevant", "relevant"),
        "suggested_enforcement_related": _string_value(
            normalized,
            "suggested_enforcement_related",
            "enforcement_related",
        ),
        "suggested_index_relevant": _string_value(
            normalized,
            "suggested_index_relevant",
            "index_relevant",
        ),
        "suggested_taxonomy_version": _string_value(
            normalized, "suggested_taxonomy_version", "taxonomy_version"
        ),
        "suggested_taxonomy_category_id": _string_value(
            normalized,
            "suggested_taxonomy_category_id",
            "taxonomy_category_id",
            "category_id",
            "tfht_category_id",
        ),
        "suggested_taxonomy_subcategory_id": _string_value(
            normalized,
            "suggested_taxonomy_subcategory_id",
            "taxonomy_subcategory_id",
            "subcategory_id",
            "tfht_subcategory_id",
        ),
        "suggested_category": _string_value(normalized, "suggested_category", "category"),
        "suggested_sub_category": _string_value(
            normalized, "suggested_sub_category", "sub_category", "subcategory"
        ),
        "suggested_confidence": _string_value(normalized, "suggested_confidence") or "manual",
        "relevant": _string_value(normalized, "relevant"),
        "enforcement_related": _string_value(normalized, "enforcement_related"),
        "index_relevant": _string_value(normalized, "index_relevant"),
        "taxonomy_version": _string_value(normalized, "taxonomy_version"),
        "taxonomy_category_id": _string_value(
            normalized,
            "taxonomy_category_id",
            "category_id",
            "tfht_category_id",
        ),
        "taxonomy_subcategory_id": _string_value(
            normalized,
            "taxonomy_subcategory_id",
            "subcategory_id",
            "tfht_subcategory_id",
        ),
        "category": _string_value(normalized, "category"),
        "sub_category": _string_value(normalized, "sub_category", "subcategory"),
        "review_status": _string_value(normalized, "review_status") or "reviewed",
        "annotation_source": _string_value(normalized, "annotation_source") or "manual_table",
        "expected_month_bucket": _string_value(normalized, "expected_month_bucket"),
        "expected_city": _string_value(normalized, "expected_city"),
        "expected_status": _string_value(normalized, "expected_status"),
        "manual_city": _string_value(normalized, "manual_city"),
        "manual_address": _string_value(normalized, "manual_address"),
        "manual_event_label": _string_value(normalized, "manual_event_label"),
        "manual_status": _string_value(normalized, "manual_status"),
        "annotation_notes": _string_value(normalized, "annotation_notes", "notes"),
        "collected_at": _string_value(normalized, "collected_at") or collected_at.isoformat(),
    }
    validate_reviewed_row(draft_row, draft_source="reviewed_examples_import")
    return draft_row


def _manual_tracking_rows(input_path: Path) -> tuple[list[ValidationDraftRow], list[str]]:
    taxonomy = default_taxonomy()
    workbook = _load_workbook(input_path)
    if "מדד האכיפה" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the 'מדד האכיפה' sheet")
    sheet = workbook["מדד האכיפה"]

    year = None
    for row_number in range(1, 5):
        for cell in sheet[row_number]:
            if isinstance(cell.value, int) and cell.value > 2000:
                year = cell.value
                break
        if year is not None:
            break
    if year is None:
        raise ValueError("Could not determine year from the manual tracking workbook")

    month_row = 3
    header_row = 4
    blocks: list[tuple[int, int]] = []
    for column in range(1, sheet.max_column + 1):
        if str(sheet.cell(header_row, column).value or "").strip() == "תאריך":
            month_name = str(sheet.cell(month_row, column).value or "").strip()
            if month_name in _MONTH_NAMES:
                blocks.append((column, _MONTH_NAMES[month_name]))

    rows: list[ValidationDraftRow] = []
    warnings: list[str] = []
    collected_at = datetime.now(UTC)

    for row_index in range(header_row + 1, sheet.max_row + 1):
        for start_column, month in blocks:
            raw_day = sheet.cell(row_index, start_column).value
            address = str(sheet.cell(row_index, start_column + 1).value or "").strip()
            event_label = str(sheet.cell(row_index, start_column + 2).value or "").strip()
            relevant_details = str(sheet.cell(row_index, start_column + 3).value or "").strip()
            status_text = str(sheet.cell(row_index, start_column + 4).value or "").strip()
            source_info = str(sheet.cell(row_index, start_column + 5).value or "").strip()

            if not any((raw_day, address, event_label, relevant_details, status_text, source_info)):
                continue

            urls = _extract_urls(source_info)
            if not urls:
                warnings.append(
                    f"row {row_index} month {month}: skipped because no source URL was found"
                )
                continue

            inferred = _infer_taxonomy_leaf(event_label, relevant_details, status_text)
            if inferred is None:
                warnings.append(
                    f"row {row_index} month {month}: skipped because taxonomy inference failed"
                )
                continue

            day = _parse_day(raw_day)
            if day is None:
                warnings.append(
                    f"row {row_index} month {month}: skipped because date parsing failed"
                )
                continue

            taxonomy_category_id, taxonomy_subcategory_id = inferred
            legacy_category, legacy_sub_category = taxonomy.legacy_mapping(
                taxonomy_category_id,
                taxonomy_subcategory_id,
            )
            first_url = urls[0].rstrip(");")
            canonical_url = canonicalize_news_url(first_url)
            title_parts = [event_label]
            if address:
                title_parts.append(address)
            rows.append(
                ValidationDraftRow(
                    source_name=_source_name_for_url(first_url),
                    article_date=datetime(year, month, day, tzinfo=UTC),
                    url=first_url,
                    canonical_url=canonical_url,
                    title=" - ".join(part for part in title_parts if part),
                    snippet=relevant_details or event_label,
                    suggested_relevant=True,
                    suggested_enforcement_related=True,
                    suggested_index_relevant=taxonomy.is_index_relevant(
                        taxonomy_category_id,
                        taxonomy_subcategory_id,
                    ),
                    suggested_taxonomy_version=taxonomy.version,
                    suggested_taxonomy_category_id=taxonomy_category_id,
                    suggested_taxonomy_subcategory_id=taxonomy_subcategory_id,
                    suggested_category=legacy_category.value,
                    suggested_sub_category=legacy_sub_category.value if legacy_sub_category else "",
                    suggested_confidence="high",
                    relevant=True,
                    enforcement_related=True,
                    index_relevant=taxonomy.is_index_relevant(
                        taxonomy_category_id,
                        taxonomy_subcategory_id,
                    ),
                    taxonomy_version=taxonomy.version,
                    taxonomy_category_id=taxonomy_category_id,
                    taxonomy_subcategory_id=taxonomy_subcategory_id,
                    category=legacy_category.value,
                    sub_category=legacy_sub_category.value if legacy_sub_category else "",
                    review_status="reviewed",
                    annotation_source=TFHT_MANUAL_TRACKING_V1,
                    manual_city=_infer_city(address),
                    manual_address=address,
                    manual_event_label=event_label,
                    manual_status=status_text,
                    annotation_notes="Imported from TFHT manual tracking workbook",
                    collected_at=collected_at,
                )
            )

    deduped: dict[tuple[str, str], ValidationDraftRow] = {}
    for draft_row in rows:
        deduped[(draft_row.source_name, draft_row.canonical_url)] = draft_row
    skipped_duplicates = len(rows) - len(deduped)
    if skipped_duplicates:
        warnings.append(f"Skipped {skipped_duplicates} duplicate row(s) by source/canonical_url")
    return list(deduped.values()), warnings


def _reviewed_examples_rows(input_path: Path) -> tuple[list[ValidationDraftRow], list[str]]:
    raw_rows = _read_tabular_rows(input_path)
    collected_at = datetime.now(UTC)
    rows: list[ValidationDraftRow] = []
    warnings: list[str] = []

    for row_number, raw_row in enumerate(raw_rows, start=2):
        try:
            normalized = _normalize_reviewed_examples_row(raw_row, collected_at=collected_at)
            rows.append(ValidationDraftRow.model_validate(normalized))
        except (ValueError, ValidationError) as exc:
            warnings.append(f"row {row_number}: skipped because {exc}")
        except Exception as exc:
            msg = f"row {row_number}: unexpected import failure"
            raise RuntimeError(msg) from exc

    deduped: dict[tuple[str, str], ValidationDraftRow] = {}
    for draft_row in rows:
        deduped[(draft_row.source_name, draft_row.canonical_url)] = draft_row
    skipped_duplicates = len(rows) - len(deduped)
    if skipped_duplicates:
        warnings.append(f"Skipped {skipped_duplicates} duplicate row(s) by source/canonical_url")
    return list(deduped.values()), warnings


def import_reviewed_table(
    *,
    input_path: Path,
    format_name: str,
    output_path: Path | None = None,
) -> ReviewedTableImportResult:
    """Import a reviewed external table into the validation draft CSV shape."""
    if not input_path.exists():
        raise FileNotFoundError(f"Reviewed table not found: {input_path}")

    if format_name == TFHT_MANUAL_TRACKING_V1:
        rows, warnings = _manual_tracking_rows(input_path)
    elif format_name == VALIDATION_REVIEWED_EXAMPLES_V1:
        rows, warnings = _reviewed_examples_rows(input_path)
    else:
        raise ValueError(f"Unsupported reviewed-table format: {format_name}")

    final_output_path = output_path or _default_output_path(input_path, format_name)
    write_csv_rows(final_output_path, DRAFT_COLUMNS, serialize_draft_rows(rows))
    skipped_rows = sum(1 for warning in warnings if "skipped because" in warning)
    return ReviewedTableImportResult(
        output_path=final_output_path,
        imported_rows=len(rows),
        skipped_rows=skipped_rows,
        warnings=warnings,
    )
