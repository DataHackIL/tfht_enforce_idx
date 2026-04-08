"""Tests for reviewed-table import into validation drafts."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from denbust.validation.common import read_csv_rows
from denbust.validation.import_reviewed import (
    TFHT_MANUAL_TRACKING_V1,
    VALIDATION_REVIEWED_EXAMPLES_V1,
    _default_output_path,
    _infer_city,
    _infer_taxonomy_leaf,
    _parse_day,
    _source_name_for_url,
    import_reviewed_table,
)


def _build_manual_tracking_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "מדד האכיפה"
    sheet["B2"] = 2026
    sheet["B3"] = "ינואר"
    headers = ["תאריך", "כתובת", "אירוע", "פרטים רלוונטיים", "סטטוס", "מקור מידע"]
    for index, value in enumerate(headers, start=2):
        sheet.cell(4, index, value)

    sheet["B5"] = 8
    sheet["C5"] = "בני ברק"
    sheet["D5"] = "מעצר חשודה על החזקת מקום לשם זנות"
    sheet["E5"] = "עינת הראל נעצרה; אותר בית בושת פעיל"
    sheet["F5"] = "פתוח"
    sheet["G5"] = (
        "https://www.maariv.co.il/news/law/article-1270778; "
        "https://www.mako.co.il/men-men_news/Article-326411394f9fb91026.htm?Partner=searchResults"
    )

    sheet["B6"] = 9
    sheet["C6"] = "אשקלון"
    sheet["D6"] = "סגירה מנהלית"
    sheet["E6"] = "צו מנהלי לדירה ברחוב ביאליק"
    sheet["F6"] = "סגור עד 23.2.26"
    sheet["G6"] = "https://www.kan-ashkelon.co.il/news/100735"

    sheet["B7"] = 10
    sheet["C7"] = "תל אביב"
    sheet["D7"] = "מעצר חשודה"
    sheet["E7"] = "אין מידע מספק"
    sheet["F7"] = ""
    sheet["G7"] = ""

    sheet["B8"] = 8
    sheet["C8"] = "בני ברק"
    sheet["D8"] = "מעצר חשודה על החזקת מקום לשם זנות"
    sheet["E8"] = "עינת הראל נעצרה; אותר בית בושת פעיל"
    sheet["F8"] = "פתוח"
    sheet["G8"] = "https://www.maariv.co.il/news/law/article-1270778"

    workbook.save(path)


def test_import_reviewed_table_normalizes_manual_tracking_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "manual_tracking.xlsx"
    _build_manual_tracking_workbook(workbook_path)

    result = import_reviewed_table(
        input_path=workbook_path,
        format_name=TFHT_MANUAL_TRACKING_V1,
    )

    rows = read_csv_rows(result.output_path)
    assert result.imported_rows == 2
    assert result.skipped_rows == 1
    assert any("duplicate" in warning for warning in result.warnings)

    first_row = rows[0]
    second_row = rows[1]
    assert first_row["review_status"] == "reviewed"
    assert first_row["taxonomy_category_id"] == "brothels"
    assert first_row["taxonomy_subcategory_id"] == "keeping_brothel"
    assert first_row["category"] == "brothel"
    assert first_row["source_name"] == "maariv"
    assert first_row["canonical_url"] == "https://maariv.co.il/news/law/article-1270778"
    assert first_row["annotation_source"] == TFHT_MANUAL_TRACKING_V1

    assert second_row["taxonomy_subcategory_id"] == "administrative_closure"
    assert second_row["index_relevant"] == "True"


def test_import_reviewed_table_rejects_unknown_format(tmp_path: Path) -> None:
    workbook_path = tmp_path / "manual_tracking.xlsx"
    _build_manual_tracking_workbook(workbook_path)

    try:
        import_reviewed_table(input_path=workbook_path, format_name="unknown")
    except ValueError as error:
        assert "Unsupported reviewed-table format" in str(error)
    else:
        raise AssertionError("Expected unsupported format to raise ValueError")


def test_import_reviewed_table_rejects_missing_file(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"

    with pytest.raises(FileNotFoundError, match="Reviewed table not found"):
        import_reviewed_table(input_path=missing, format_name=TFHT_MANUAL_TRACKING_V1)


def test_helper_functions_cover_parsing_and_taxonomy_inference() -> None:
    assert _default_output_path(Path("/tmp/manual.xlsx"), TFHT_MANUAL_TRACKING_V1).name == (
        "manual_tfht_manual_tracking_v1_draft.csv"
    )
    assert _source_name_for_url("https://www.maariv.co.il/news/law/article-1") == "maariv"
    assert _infer_city("רחוב ביאליק 1, אשקלון") == "אשקלון"
    assert _infer_city("") == ""

    assert _parse_day(None) is None
    assert _parse_day("") is None
    assert _parse_day(8) == 8
    assert _parse_day(8.0) == 8
    assert _parse_day("  ") is None
    assert _parse_day("12/03") == 12
    assert _parse_day("יום שני") is None

    assert _infer_taxonomy_leaf("ערעור על צו", "", "") == ("brothels", "closure_appeal")
    assert _infer_taxonomy_leaf("סגירה מנהלית", "", "") == (
        "brothels",
        "administrative_closure",
    )
    assert _infer_taxonomy_leaf("סחר בבני אדם", "עבדות", "") == (
        "human_trafficking",
        "trafficking_slavery_conditions",
    )
    assert _infer_taxonomy_leaf("סחר בבני אדם", "נשים מברזיל הוסגר", "") == (
        "human_trafficking",
        "trafficking_cross_border_prostitution",
    )
    assert _infer_taxonomy_leaf("סחר בנשים", "", "") == ("human_trafficking", "trafficking_women")
    assert _infer_taxonomy_leaf("סחר בבני אדם", "", "") == (
        "human_trafficking",
        "trafficking_sexual_exploitation",
    )
    assert _infer_taxonomy_leaf("כתב אישום", "בית בושת", "") == (
        "brothels",
        "brothel_indictment",
    )
    assert _infer_taxonomy_leaf("קנס לזנאי", "", "") == ("brothels", "client_fine")
    assert _infer_taxonomy_leaf("שידול לזנות", "", "") == (
        "pimping_prostitution",
        "soliciting_prostitution",
    )
    assert _infer_taxonomy_leaf("הבאת אדם לידי זנות", "", "") == (
        "pimping_prostitution",
        "bringing_into_prostitution",
    )
    assert _infer_taxonomy_leaf("סרסרות", "", "") == ("pimping_prostitution", "pimping")
    assert _infer_taxonomy_leaf("זנות מקוונת", "", "") == (
        "pimping_prostitution",
        "online_prostitution",
    )
    assert _infer_taxonomy_leaf("השכרת מקום", "", "") == ("brothels", "renting_brothel")
    assert _infer_taxonomy_leaf("פרסום", "זנות", "") == ("brothels", "advertising_prostitution")
    assert _infer_taxonomy_leaf("דירה דיסקרטית", "", "") == ("brothels", "keeping_brothel")
    assert _infer_taxonomy_leaf("מעצר חשודה", "אין מידע מספק", "") is None


def test_import_reviewed_table_rejects_missing_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "not-it"
    workbook_path = tmp_path / "missing_sheet.xlsx"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="missing the 'מדד האכיפה' sheet"):
        import_reviewed_table(input_path=workbook_path, format_name=TFHT_MANUAL_TRACKING_V1)


def test_import_reviewed_table_rejects_missing_year(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "מדד האכיפה"
    sheet["B3"] = "ינואר"
    sheet["B4"] = "תאריך"
    workbook_path = tmp_path / "missing_year.xlsx"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="Could not determine year"):
        import_reviewed_table(input_path=workbook_path, format_name=TFHT_MANUAL_TRACKING_V1)


def test_import_reviewed_table_skips_rows_for_missing_url_bad_taxonomy_and_bad_date(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "מדד האכיפה"
    sheet["B2"] = 2026
    sheet["B3"] = "ינואר"
    for index, value in enumerate(
        ["תאריך", "כתובת", "אירוע", "פרטים רלוונטיים", "סטטוס", "מקור מידע"], start=2
    ):
        sheet.cell(4, index, value)

    sheet["B5"] = 1
    sheet["C5"] = "תל אביב"
    sheet["D5"] = "סרסרות"
    sheet["E5"] = "פרטים"
    sheet["F5"] = ""
    sheet["G5"] = ""

    sheet["B6"] = 2
    sheet["C6"] = "ירושלים"
    sheet["D6"] = "מעצר חשודה"
    sheet["E6"] = "אין מידע מספק"
    sheet["F6"] = ""
    sheet["G6"] = "https://example.com/a"

    sheet["B7"] = "יום שני"
    sheet["C7"] = "חיפה"
    sheet["D7"] = "סרסרות"
    sheet["E7"] = "פרטים"
    sheet["F7"] = ""
    sheet["G7"] = "https://example.com/b"

    workbook_path = tmp_path / "skips.xlsx"
    workbook.save(workbook_path)

    result = import_reviewed_table(input_path=workbook_path, format_name=TFHT_MANUAL_TRACKING_V1)

    assert result.imported_rows == 0
    assert result.skipped_rows == 3
    assert any("no source URL was found" in warning for warning in result.warnings)
    assert any("taxonomy inference failed" in warning for warning in result.warnings)
    assert any("date parsing failed" in warning for warning in result.warnings)


def test_import_reviewed_table_normalizes_generic_reviewed_examples_csv(tmp_path: Path) -> None:
    input_path = tmp_path / "reviewed_examples.csv"
    input_path.write_text(
        "\n".join(
            [
                (
                    "URL,Title,Snippet,Article Date,Source Name,Relevant,Enforcement Related,"
                    "Index Relevant,Taxonomy Version,Taxonomy Category ID,Taxonomy Subcategory ID,"
                    "Category,Sub Category,Review Status,Annotation Source,Expected Month Bucket,"
                    "Expected City,Expected Status"
                ),
                (
                    "https://example.com/keep?utm_source=one,כתבה א,סיכום א,2026-04-01T00:00:00+00:00,"
                    "ynet,True,True,True,1,brothels,administrative_closure,brothel,closure,"
                    "reviewed,manual table,2026-04,חיפה,closed"
                ),
                (
                    "https://example.com/keep?utm_source=two,כתבה א,סיכום א,2026-04-01T00:00:00+00:00,"
                    "ynet,True,True,True,1,brothels,administrative_closure,brothel,closure,"
                    "reviewed,manual table,2026-04,חיפה,closed"
                ),
                (
                    "https://example.com/outside,כתבה ב,סיכום ב,2026-04-02T00:00:00+00:00,"
                    ",False,False,False,,,,not_relevant,,reviewed,Google alerts comparison,,,"
                ),
            ]
        ),
        encoding="utf-8",
    )

    result = import_reviewed_table(
        input_path=input_path,
        format_name=VALIDATION_REVIEWED_EXAMPLES_V1,
    )

    rows = read_csv_rows(result.output_path)
    assert result.imported_rows == 2
    assert result.skipped_rows == 0
    assert any("duplicate" in warning for warning in result.warnings)

    first_row = rows[0]
    second_row = rows[1]
    assert first_row["canonical_url"] == "https://example.com/keep"
    assert first_row["annotation_source"] == "manual table"
    assert first_row["expected_month_bucket"] == "2026-04"
    assert first_row["expected_city"] == "חיפה"
    assert first_row["expected_status"] == "closed"
    assert second_row["source_name"] == "example"
    assert second_row["category"] == "not_relevant"
    assert second_row["annotation_source"] == "Google alerts comparison"


def test_import_reviewed_table_normalizes_generic_reviewed_examples_xlsx(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "URL"
    sheet["B1"] = "Title"
    sheet["C1"] = "Snippet"
    sheet["D1"] = "Article Date"
    sheet["E1"] = "Relevant"
    sheet["F1"] = "Enforcement Related"
    sheet["G1"] = "Index Relevant"
    sheet["H1"] = "Taxonomy Category ID"
    sheet["I1"] = "Taxonomy Subcategory ID"
    sheet["J1"] = "Category"
    sheet["K1"] = "Sub Category"
    sheet["A2"] = "https://example.com/xlsx"
    sheet["B2"] = "כותרת"
    sheet["C2"] = "תקציר"
    sheet["D2"] = "2026-04-03T00:00:00+00:00"
    sheet["E2"] = True
    sheet["F2"] = True
    sheet["G2"] = True
    sheet["H2"] = "brothels"
    sheet["I2"] = "administrative_closure"
    sheet["J2"] = "brothel"
    sheet["K2"] = "closure"
    sheet["A3"] = "https://example.com/xlsx-false"
    sheet["B3"] = "כותרת שלילית"
    sheet["C3"] = "תקציר שלילי"
    sheet["D3"] = "2026-04-04T00:00:00+00:00"
    sheet["E3"] = False
    sheet["F3"] = False
    sheet["G3"] = False
    sheet["J3"] = "not_relevant"
    workbook_path = tmp_path / "reviewed_examples.xlsx"
    workbook.save(workbook_path)

    result = import_reviewed_table(
        input_path=workbook_path,
        format_name=VALIDATION_REVIEWED_EXAMPLES_V1,
    )

    rows = read_csv_rows(result.output_path)
    assert result.imported_rows == 2
    assert rows[0]["canonical_url"] == "https://example.com/xlsx"
    assert rows[0]["source_name"] == "example"
    assert rows[1]["canonical_url"] == "https://example.com/xlsx-false"
    assert rows[1]["relevant"] == "False"
    assert rows[1]["enforcement_related"] == "False"
    assert rows[1]["index_relevant"] == "False"
    assert rows[1]["category"] == "not_relevant"


def test_import_reviewed_table_generic_adapter_rejects_invalid_reviewed_labels(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "reviewed_examples.csv"
    input_path.write_text(
        "\n".join(
            [
                (
                    "URL,Title,Article Date,Relevant,Enforcement Related,Index Relevant,"
                    "Taxonomy Version,Taxonomy Category ID,Taxonomy Subcategory ID,Category,"
                    "Sub Category,Review Status"
                ),
                (
                    "https://example.com/bad,כתבה,2026-04-01T00:00:00+00:00,True,True,False,"
                    "1,brothels,administrative_closure,brothel,closure,reviewed"
                ),
                "https://example.com/missing-title,,2026-04-01T00:00:00+00:00,True,False,False,,,,brothel,,reviewed",
            ]
        ),
        encoding="utf-8",
    )

    result = import_reviewed_table(
        input_path=input_path,
        format_name=VALIDATION_REVIEWED_EXAMPLES_V1,
    )

    assert result.imported_rows == 0
    assert result.skipped_rows == 2
    assert any(
        "index_relevant does not match the packaged taxonomy" in warning
        for warning in result.warnings
    )
    assert any("title is required" in warning for warning in result.warnings)


def test_import_reviewed_table_generic_adapter_handles_empty_xlsx(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook_path = tmp_path / "empty_reviewed_examples.xlsx"
    workbook.save(workbook_path)

    result = import_reviewed_table(
        input_path=workbook_path,
        format_name=VALIDATION_REVIEWED_EXAMPLES_V1,
    )

    rows = read_csv_rows(result.output_path)
    assert result.imported_rows == 0
    assert result.skipped_rows == 0
    assert result.warnings == []
    assert rows == []


def test_import_reviewed_table_generic_adapter_rejects_unsupported_file_type(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "reviewed_examples.txt"
    input_path.write_text("placeholder", encoding="utf-8")

    with pytest.raises(ValueError, match="Unsupported reviewed examples file type"):
        import_reviewed_table(
            input_path=input_path,
            format_name=VALIDATION_REVIEWED_EXAMPLES_V1,
        )


def test_import_reviewed_table_generic_adapter_requires_url_and_article_date(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "reviewed_examples_missing_fields.csv"
    input_path.write_text(
        "\n".join(
            [
                "URL,Title,Article Date,Relevant,Enforcement Related,Category,Review Status",
                ",כתבה ללא URL,2026-04-01T00:00:00+00:00,False,False,not_relevant,reviewed",
                "https://example.com/missing-date,כתבה ללא תאריך,,False,False,not_relevant,reviewed",
            ]
        ),
        encoding="utf-8",
    )

    result = import_reviewed_table(
        input_path=input_path,
        format_name=VALIDATION_REVIEWED_EXAMPLES_V1,
    )

    assert result.imported_rows == 0
    assert result.skipped_rows == 2
    assert any("url is required" in warning for warning in result.warnings)
    assert any("article_date is required" in warning for warning in result.warnings)
